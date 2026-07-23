# -*- coding: utf-8 -*-
"""광안 보안등 유선 0-10V 디밍 조명제어 견적서 (편집형 xlsx)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "견적서"

NAVY = "12315E"; BLUE = "1E40AF"; LIGHT = "EFF4FB"; GRAY = "6B7280"; SEC = "DCE6F5"
thin = Side(style="thin", color="B8C2D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wfont = Font(name="맑은 고딕", color="FFFFFF", bold=True, size=10)
hfont = Font(name="맑은 고딕", bold=True, size=10)
nfont = Font(name="맑은 고딕", size=9.5)
sfont = Font(name="맑은 고딕", bold=True, size=9.5, color=NAVY)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
right = Alignment(horizontal="right", vertical="center")

# columns: No | 품명 | 규격 | 단위 | 수량 | 단가 | 금액 | 비고
widths = [5, 30, 34, 6, 6, 13, 15, 22]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 1
# Title
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = ws.cell(r, 1, "견  적  서")
c.font = Font(name="맑은 고딕", bold=True, size=20, color=NAVY)
c.alignment = center
ws.row_dimensions[r].height = 34
r += 2

# Header info block
info = [
    ("공 사 명", "광안 지역주택조합 공동주택 신축공사 — 옥외 보안등 조명제어 시스템", "제 출 처", "삼환전기 귀중"),
    ("현    장", "부산 수영구 광안동 971", "제 출 일", "2026. 07. 24."),
    ("제어대상", "옥외 보안등 41등 / 6회로 (LCP-0, 501동 경비실)", "문    서", "v1.0 (E13 상세도 반영)"),
    ("공급자", "UTTEC (유티이씨) · 대표이사 홍광선", "연 락 처", "010-2401-9059 / ihong9059@gmail.com"),
    ("주    소", "경기도 용인시 기흥구 흥덕중앙로 120 흥덕유타워 2404호", "웹", "www.uttec.co.kr"),
]
for a, b, cc, d in info:
    ws.cell(r, 1, a).font = sfont; ws.cell(r, 1).fill = PatternFill("solid", fgColor=LIGHT); ws.cell(r,1).alignment=center
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r, 2, b).font = nfont; ws.cell(r, 2).alignment = left
    ws.cell(r, 6, cc).font = sfont; ws.cell(r, 6).fill = PatternFill("solid", fgColor=LIGHT); ws.cell(r,6).alignment=center
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    ws.cell(r, 7, d).font = nfont; ws.cell(r, 7).alignment = left
    for col in range(1, 9):
        ws.cell(r, col).border = border
    ws.row_dimensions[r].height = 22
    r += 1
r += 1

# Column header
heads = ["No", "품    명", "규    격", "단위", "수량", "단가(원)", "금액(원)", "비    고"]
for i, h in enumerate(heads, 1):
    c = ws.cell(r, i, h); c.font = wfont; c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = center; c.border = border
ws.row_dimensions[r].height = 24
header_row = r
r += 1

# line items: (section?, name, spec, unit, qty, unit_price, remark)
DATA = [
    ("SEC", "1. 조명제어 기기 (Control Devices)"),
    (None, "중앙전송장치 SCU", "32bit MCU, RS485 ELC+Ethernet+MODBUS, WIFI, 130×130, AC220/110V", "EA", 1, 1800000, "E13-011 사양"),
    (None, "0-10V 디밍모듈 (8채널)", "0-10V 디밍 8ch + 16A 래칭릴레이 8ch, 0~254단계", "EA", 1, 1200000, "6회로 사용+2예비"),
    (None, "전원공급장치 (Power Supply Unit)", "AC220V→24VAC, 60Hz, 60VA (Expansion Power 포함)", "EA", 1, 350000, ""),
    ("SEC", "2. 조명제어반 LCP-0 (Panel)"),
    (None, "제어반 외함 (Enclosure)", "STEEL, 400×1600×180, 자립/벽부형, 도장·명판", "EA", 1, 1500000, "E13-012 일람"),
    (None, "속판 및 내부 마운팅 자재", "속판, 덕트, 레일, 명판", "식", 1, 350000, ""),
    (None, "주차단기·회로차단기", "주 MCCB 1 + 회로별 ELB (6회로+예비)", "식", 1, 700000, ""),
    (None, "단자대·배선자재·부속", "터미널블록, 제어배선, 압착단자, 마그네트 등", "식", 1, 550000, ""),
    (None, "제어반 제작·조립·내부결선", "기기 취부, 내부 결선, 검사", "식", 1, 900000, ""),
    ("SEC", "3. 중앙감시반 CCMS (501동 경비실)"),
    (None, "관제용 PC", "Intel i5, RAM 32GB, HDD 1TB, Windows 10", "EA", 1, 1600000, "E13-011 사양"),
    (None, "24″ LED 모니터", "1920×1080, High Color", "EA", 1, 250000, ""),
    (None, "프린터", "경보·리포트 출력용", "EA", 1, 250000, ""),
    (None, "HMI 소프트웨어", "조명제어 HMI, GS 인증+BTL 인증, 스케쥴·경보·그래픽·경향분석·자기진단", "식", 1, 4500000, "인증 제품"),
    (None, "관제 콘솔 데스크", "경비실 설치대 (옵션)", "식", 1, 400000, "옵션"),
    ("SEC", "4. 통신·제어 자재 (결선용 / 배관·포설은 전기공사 별도)"),
    (None, "데이터라인 케이블", "F-CVVS 1.5㎟ × 2C (SCU↔LCP, 22C)", "식", 1, 350000, "약 200m 기준"),
    (None, "0-10V 디밍 통신선", "AWG18 (UL2095, 2C), 16C", "식", 1, 300000, "안정기 구간"),
    (None, "제어반 전원선", "F-CV 2.5㎟ × 3C / HFIX 2.5㎟", "식", 1, 200000, ""),
    (None, "단자·성단·소모 자재", "압착단자, 케이블타이, 마킹 등", "식", 1, 200000, ""),
    ("SEC", "5. 설치·용역 (Installation & Service)"),
    (None, "조명제어 기기 셋팅·프로그래밍", "SCU/디밍모듈/전원 파라미터, 그룹·씬·스케쥴(41등/6회로)", "식", 1, 1500000, ""),
    (None, "기기 신호선 결선", "DIM1~8 / R1~8 / ETLC LINE / 24VAC 결선", "식", 1, 800000, ""),
    (None, "0-10V 통신라인 결선", "디밍모듈↔안정기 결선·성단", "식", 1, 600000, ""),
    (None, "CCMS HMI 그래픽 작성·포인트 설정", "단지 배치도 그래픽, 포인트 매핑, 경보·로깅 설정", "식", 1, 1800000, ""),
    (None, "시스템 통합시험·시운전(T&C)", "회로별 조광·스케쥴 검증, 통합 시운전", "식", 1, 1200000, ""),
    (None, "운영자 교육", "경비실 운영자 사용 교육, 매뉴얼 제공", "식", 1, 300000, ""),
    ("SEC", "6. 기타 (Others)"),
    (None, "준공도서·시험성적서·인증서", "준공도서, 시험성적서, 인증서 사본", "식", 1, 300000, ""),
    (None, "운반비", "기기·제어반 부산 현장 운송", "식", 1, 400000, ""),
    (None, "출장·숙박비", "시공·시운전 출장 (부산)", "식", 1, 600000, ""),
    (None, "예비품 (Spare)", "디밍 예비 2회로분·퓨즈·소모품", "식", 1, 400000, ""),
]

no = 0
amount_rows = []
for row in DATA:
    if row[0] == "SEC":
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        c = ws.cell(r, 1, row[1]); c.font = Font(name="맑은 고딕", bold=True, size=10, color=NAVY)
        c.fill = PatternFill("solid", fgColor=SEC); c.alignment = Alignment(horizontal="left", vertical="center")
        for col in range(1, 9):
            ws.cell(r, col).border = border
        ws.row_dimensions[r].height = 20
        r += 1
        continue
    no += 1
    _, name, spec, unit, qty, price, remark = row
    ws.cell(r, 1, no).alignment = center; ws.cell(r, 1).font = nfont
    ws.cell(r, 2, name).alignment = left; ws.cell(r, 2).font = Font(name="맑은 고딕", size=9.5, bold=True)
    ws.cell(r, 3, spec).alignment = left; ws.cell(r, 3).font = nfont
    ws.cell(r, 4, unit).alignment = center; ws.cell(r, 4).font = nfont
    ws.cell(r, 5, qty).alignment = center; ws.cell(r, 5).font = nfont
    ws.cell(r, 6, price).alignment = right; ws.cell(r, 6).font = nfont; ws.cell(r, 6).number_format = "#,##0"
    ws.cell(r, 7, f"=E{r}*F{r}").alignment = right; ws.cell(r, 7).font = nfont; ws.cell(r, 7).number_format = "#,##0"
    ws.cell(r, 8, remark).alignment = left; ws.cell(r, 8).font = Font(name="맑은 고딕", size=8.5, color=GRAY)
    for col in range(1, 9):
        ws.cell(r, col).border = border
    ws.row_dimensions[r].height = 26
    amount_rows.append(r)
    r += 1

first_amt, last_amt = amount_rows[0], amount_rows[-1]

def total_row(label, formula, bold=True, fill=None):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c = ws.cell(r, 1, label)
    c.font = Font(name="맑은 고딕", bold=bold, size=10, color=NAVY)
    c.alignment = Alignment(horizontal="right", vertical="center")
    a = ws.cell(r, 7, formula); a.number_format = "#,##0"
    a.font = Font(name="맑은 고딕", bold=bold, size=10, color=NAVY)
    a.alignment = right
    ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=8)
    if fill:
        for col in range(1, 9):
            ws.cell(r, col).fill = PatternFill("solid", fgColor=fill)
    for col in range(1, 9):
        ws.cell(r, col).border = border
    ws.row_dimensions[r].height = 22
    r += 1

sub_row = r
total_row("직접비 소계 (①)", f"=SUM(G{first_amt}:G{last_amt})", fill=LIGHT)
gm_row = r
total_row("일반관리비 (② = ① × 5%)", f"=ROUND(G{sub_row}*0.05,0)")
profit_row = r
total_row("이윤 (③ = (①+②) × 10%)", f"=ROUND((G{sub_row}+G{gm_row})*0.10,0)")
supply_row = r
total_row("공급가액 (④ = ①+②+③)", f"=G{sub_row}+G{gm_row}+G{profit_row}", fill=LIGHT)
vat_row = r
total_row("부가가치세 (⑤ = ④ × 10%)", f"=ROUND(G{supply_row}*0.10,0)")
total_row("총 합계 금액 (④+⑤, VAT 포함)", f"=G{supply_row}+G{vat_row}", fill="FDE68A")
r += 1

# notes
notes = [
    "※ 견적 조건",
    "1. 본 견적은 「E13-011~013 보안등 조명제어 상세도」 기준 UTTEC(조명제어공사) 범위 산출입니다.",
    "2. 제외 범위: 신호선 배관·전선 포설·전원공사(전기공사), 등기구 및 0-10V 조광 안정기(등기구 업체).",
    "3. 제어반 외함 SIZE(400×1600×180)는 현장 여건에 따라 변경될 수 있으며, 이 경우 금액 조정됩니다.",
    "4. 등기구 조광 방식(0-10V / DALI) 확정 및 현장 실사 후 최종 금액이 확정됩니다.",
    "5. 상기 단가는 제안 기준가이며, 세부 사양·수량 확정 및 발주 조건에 따라 협의 조정 가능합니다.",
    "6. 유효기간: 견적일로부터 30일. 결제 조건 및 하자보증(1년)은 계약 시 협의합니다.",
]
for i, t in enumerate(notes):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = ws.cell(r, 1, t)
    c.font = Font(name="맑은 고딕", bold=(i == 0), size=9 if i == 0 else 8.7, color=NAVY if i == 0 else "374151")
    c.alignment = left
    ws.row_dimensions[r].height = 18
    r += 1

ws.print_options.horizontalCentered = True
ws.page_setup.orientation = "portrait"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

out = "삼환_광안보안등_유선디밍제어_견적서_UTTEC_v1_20260724.xlsx"
wb.save(out)
print("saved", out, "| line items:", no)
