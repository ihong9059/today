# -*- coding: utf-8 -*-
"""[내부용] 광안 보안등 조명제어 구매참고 BOM — 실제 모델·시장가 + 마진분석
xlsx(수식 편집형) + HTML(→PDF) 동시 생성. 단일 데이터 소스."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 구분, 품목, 도면사양, 권장 제조사·모델, 시장단가(추정중간), 신뢰도, 견적단가, 구매처/비고
# 신뢰도: ● 웹실검 / ◐ 시장지식추정 / ○ 전용기기 제조사견적필요 / ▷ 용역(공임원가)
DATA = [
    ("SEC", "1. 조명제어 전용기기 (㈜엠알바스 MRBAS · ELC SYSTEM — 도면 'Excellent Lighting Controls')"),
    ("자재", "중앙전송장치 SCU", "32bit MCU, RS485 ELC+Ethernet+MODBUS, 999모듈", "MRBAS ELC 'SCU' (System Communication Unit)", 1900000, "○", 1800000, "제조사 직견적 필수 (전용)"),
    ("자재", "0-10V 디밍모듈 8ch", "0-10V 디밍 8ch, 0~254단계", "MRBAS '0-10V Dimming Module'(DCM/LDCM계열)", 1300000, "○", 1200000, "제조사 견적 / 릴레이 별도 여부 확인"),
    ("자재", "릴레이모듈 (16A 래칭 8접점)", "16A 래칭릴레이 R1~8", "MRBAS '6eRM-w/16a' + LR-9P 릴레이(플러그인)", 750000, "○", 0, "도면상 디밍모듈에 포함 표기 — 실제 별도품일 수 있음(원가확인용 분리)"),
    ("자재", "전원공급장치", "AC220V→24VAC 60Hz, 60VA", "MRBAS 'RPWR' (AC220/DC24V, 40W)", 320000, "○", 350000, "제조사 견적"),
    ("자재", "HMI 관제 소프트웨어", "조명제어 HMI, GS·BTL 인증", "MRBAS ELC 중앙관제 S/W (라이선스)", 4000000, "○", 4500000, "전용 라이선스 — 제조사 견적 필수(변동 큼)"),
    ("SEC", "2. 조명제어반 LCP-0 함체·배전 (표준 전기자재)"),
    ("자재", "제어반 외함", "STEEL 400×1600×180, 자립/벽부, 도장·명판", "탑박스/박스파라 등 제작 (STEEL 함체)", 700000, "◐", 1500000, "topbox.co.kr / boxpara.com 제작견적"),
    ("자재", "속판·덕트·레일·명판", "내부 마운팅 자재", "함체업체 부속 (속판 세트)", 220000, "◐", 350000, "함체와 동시 발주"),
    ("자재", "주차단기·회로차단기", "주MCCB 1 + 회로ELB 6~7", "LS ELECTRIC 주 ABS103c 3P + 회로 EBS53c 3P ×7", 380000, "●", 700000, "ABE33B 3P30A≈30,940 / EBS54C ELB (풍림·2sk·speedmall)"),
    ("자재", "단자대·제어배선자재", "터미널블록, 제어선, 압착단자", "진흥전기/광명전기 or Phoenix/WAGO 단자", 280000, "◐", 550000, ""),
    ("용역", "제어반 제작·조립·내부결선", "기기취부·내부결선·검사", "판넬 제작업체 공임", 650000, "▷", 900000, "함체업체 일괄 시 절감"),
    ("SEC", "3. 중앙감시 CCMS (표준 IT장비)"),
    ("자재", "관제용 PC", "i5, RAM 32GB, SSD 1TB, Win10/11", "HP ProDesk 400 G9 / Dell OptiPlex (i5-14400급)", 1900000, "◐", 1600000, "⚠ 브랜드 32GB/1TB SSD 시세 ↑ — 견적단가 상향 검토"),
    ("자재", "24″ LED 모니터", "1920×1080", "LG 24MK430H-B / 삼성 S24C310", 175000, "●", 250000, "다나와 시세"),
    ("자재", "프린터", "경보·리포트 출력", "HP LaserJet M111w / 삼성 흑백레이저", 200000, "●", 250000, ""),
    ("자재", "관제 콘솔 데스크", "경비실 설치대(옵션)", "사무용 데스크", 300000, "◐", 400000, "옵션"),
    ("SEC", "4. 통신·제어 케이블 (표준 전선)"),
    ("자재", "데이터라인 F-CVVS 1.5㎟×2C", "SCU↔LCP, 22C, ~200m", "대한전선/가온전선/코리아테크 (제어용 차폐)", 250000, "●", 350000, "≈1,200원/m × 200m (koreacable.net)"),
    ("자재", "0-10V 통신선 AWG18(UL2095,2C)", "안정기 구간, 16C", "UL2095 2C 전선", 180000, "◐", 300000, ""),
    ("자재", "제어반 전원선", "F-CV 2.5㎟×3C / HFIX 2.5㎟", "대한/LS전선", 180000, "◐", 200000, ""),
    ("자재", "단자·성단·소모자재", "압착단자·케이블타이·마킹", "일반 소모자재", 150000, "◐", 200000, ""),
    ("SEC", "5. 설치·용역 (공임 원가 — 무형)"),
    ("용역", "기기 셋팅·프로그래밍", "그룹·씬·스케쥴(41등/6회로)", "엔지니어 공수", 1000000, "▷", 1500000, "MRBAS 협조 시 조정"),
    ("용역", "기기 신호선 결선", "DIM/R/ETLC/24VAC 결선", "설치 공수", 550000, "▷", 800000, ""),
    ("용역", "0-10V 통신라인 결선", "디밍모듈↔안정기 성단", "설치 공수", 400000, "▷", 600000, ""),
    ("용역", "HMI 그래픽·포인트 설정", "배치도 그래픽·포인트·경보", "SW 엔지니어 공수", 1200000, "▷", 1800000, ""),
    ("용역", "통합시험·시운전(T&C)", "회로별 검증·시운전", "시운전 공수", 800000, "▷", 1200000, ""),
    ("용역", "운영자 교육", "경비실 교육·매뉴얼", "교육 공수", 200000, "▷", 300000, ""),
    ("SEC", "6. 기타"),
    ("자재", "준공도서·성적서·인증서", "준공도서·시험성적서 사본", "문서 제작", 200000, "▷", 300000, ""),
    ("자재", "운반비", "부산 현장 운송", "화물/택배", 350000, "◐", 400000, "부산 원거리"),
    ("자재", "출장·숙박비", "시공·시운전 출장", "출장 실비", 500000, "▷", 600000, ""),
    ("자재", "예비품(Spare)", "디밍 2회로분·퓨즈·소모품", "MRBAS 예비 모듈 등", 300000, "◐", 400000, ""),
]

# ---------- xlsx ----------
wb = openpyxl.Workbook(); ws = wb.active; ws.title = "구매참고BOM"
NAVY="12315E"; RED="B91C1C"; SEC="DCE6F5"; LIGHT="EFF4FB"; GRAY="6B7280"; YEL="FDE68A"
thin=Side(style="thin",color="B8C2D0"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(**k): return Font(name="맑은 고딕",**k)
c_ct=Alignment(horizontal="center",vertical="center",wrap_text=True)
c_lf=Alignment(horizontal="left",vertical="center",wrap_text=True)
c_rt=Alignment(horizontal="right",vertical="center")

widths=[4,6,24,26,30,13,6,13,13,10,28]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

r=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=11)
c=ws.cell(r,1,"[내부 관리용 · 대외비] 광안 보안등 조명제어 — 구매참고 BOM (실모델·시장가·마진)")
c.font=F(bold=True,size=14,color=RED); c.alignment=c_ct; ws.row_dimensions[r].height=28; r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=11)
c=ws.cell(r,1,"※ 외부 제출 금지. 견적서(대외용)의 원가·마진 검토 및 발주 참고용. 전용기기(MRBAS ELC)는 발주 전 제조사 실견적 필수. 작성 2026-07-24 / UTTEC")
c.font=F(size=9,color=GRAY); c.alignment=c_lf; ws.row_dimensions[r].height=18; r+=2

heads=["No","구분","품목","도면 사양","권장 제조사·모델","시장단가(추정)","신뢰도","견적단가","단가차이","마진%","구매처·비고"]
for i,h in enumerate(heads,1):
    cell=ws.cell(r,i,h); cell.font=F(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor=NAVY); cell.alignment=c_ct; cell.border=border
ws.row_dimensions[r].height=30; r+=1

no=0; mkt_rows=[]; q_rows=[]
for row in DATA:
    if row[0]=="SEC":
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=11)
        cell=ws.cell(r,1,row[1]); cell.font=F(bold=True,size=10,color=NAVY); cell.fill=PatternFill("solid",fgColor=SEC); cell.alignment=c_lf
        for col in range(1,12): ws.cell(r,col).border=border
        ws.row_dimensions[r].height=20; r+=1; continue
    no+=1
    gubun,name,spec,model,mkt,conf,quote,note=row
    ws.cell(r,1,no).alignment=c_ct; ws.cell(r,1).font=F(size=9)
    ws.cell(r,2,gubun).alignment=c_ct; ws.cell(r,2).font=F(size=9)
    ws.cell(r,3,name).alignment=c_lf; ws.cell(r,3).font=F(size=9,bold=True)
    ws.cell(r,4,spec).alignment=c_lf; ws.cell(r,4).font=F(size=8.5)
    ws.cell(r,5,model).alignment=c_lf; ws.cell(r,5).font=F(size=8.5,color=NAVY,bold=True)
    ws.cell(r,6,mkt).alignment=c_rt; ws.cell(r,6).font=F(size=9); ws.cell(r,6).number_format="#,##0"
    ws.cell(r,7,conf).alignment=c_ct; ws.cell(r,7).font=F(size=10)
    ws.cell(r,8,quote).alignment=c_rt; ws.cell(r,8).font=F(size=9); ws.cell(r,8).number_format="#,##0"
    ws.cell(r,9,f"=H{r}-F{r}").alignment=c_rt; ws.cell(r,9).font=F(size=9); ws.cell(r,9).number_format="#,##0;[Red]-#,##0"
    ws.cell(r,10,f'=IF(F{r}=0,"",ROUND((H{r}-F{r})/F{r}*100,0))').alignment=c_rt; ws.cell(r,10).font=F(size=9); ws.cell(r,10).number_format='0"%";[Red]-0"%"'
    ws.cell(r,11,note).alignment=c_lf; ws.cell(r,11).font=F(size=8,color=GRAY)
    for col in range(1,12): ws.cell(r,col).border=border
    ws.row_dimensions[r].height=30
    mkt_rows.append(r); q_rows.append(r); r+=1

f,l=mkt_rows[0],mkt_rows[-1]
def totrow(label,fcol=None,formula=None,fill=LIGHT):
    global r
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
    cc=ws.cell(r,1,label); cc.font=F(bold=True,size=10,color=NAVY); cc.alignment=Alignment(horizontal="right",vertical="center")
    ws.cell(r,6,f"=SUM(F{f}:F{l})").number_format="#,##0"; ws.cell(r,6).font=F(bold=True,color=NAVY)
    ws.cell(r,8,f"=SUM(H{f}:H{l})").number_format="#,##0"; ws.cell(r,8).font=F(bold=True,color=NAVY)
    ws.cell(r,9,f"=H{r}-F{r}").number_format="#,##0"; ws.cell(r,9).font=F(bold=True,color=NAVY)
    ws.cell(r,10,f"=ROUND((H{r}-F{r})/F{r}*100,0)").number_format='0"%"'; ws.cell(r,10).font=F(bold=True,color=NAVY)
    for col in range(1,12): ws.cell(r,col).fill=PatternFill("solid",fgColor=fill); ws.cell(r,col).border=border
    for col in [6,8,9,10]: ws.cell(r,col).alignment=c_rt
    ws.row_dimensions[r].height=22; r+=1

totrow("합계 — 시장/원가 추정 vs 견적 직접비",fill=YEL)
# 부연 총이익 라인
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=11)
tr=r-1
c=ws.cell(r,1,f"→ 직접비 예상 총이익 = 견적직접비 − 원가추정 (H{tr}−F{tr}). 여기에 견적서 일반관리비 5% + 이윤 10% 별도 가산됨. 전용기기(SCU·디밍·HMI) 실견적에 따라 변동 큼.")
c.font=F(size=8.5,color=GRAY); c.alignment=c_lf; ws.row_dimensions[r].height=16; r+=2

# legend
for t in [
 "신뢰도: ● 웹 실검색 확인  /  ◐ 시장지식 추정  /  ○ 전용기기(제조사 실견적 필수)  /  ▷ 용역(공임 원가추정)",
 "⚠ No.11 관제PC: 브랜드 i5·32GB·1TB SSD 시세가 견적단가보다 높음 → 견적 상향 또는 사양(16GB/SSD 512GB) 하향 검토.",
 "⚠ 전용기기(MRBAS ELC: SCU·0-10V 디밍모듈·릴레이·전원·HMI)는 추정가이며 반드시 ㈜엠알바스 실견적으로 대체할 것.",
 "출처(웹): mrbas.co.kr(ELC SYSTEM 카탈로그) · LS ELECTRIC/풍림·2sk·speedmall(차단기) · koreacable.net(F-CVVS) · HP/Dell 코리아(PC) · topbox·boxpara(함체) — 2026-07-24 검색.",
]:
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=11)
    cc=ws.cell(r,1,t); cc.font=F(size=8.5,color=(RED if t.startswith("⚠") else "374151")); cc.alignment=c_lf
    ws.row_dimensions[r].height=16; r+=1

ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=0
outx="[내부용]삼환_광안보안등_조명제어_구매참고BOM_UTTEC_v1_20260724.xlsx"
wb.save(outx)

# 합계 계산(파이썬)
mkt=sum(row[4] for row in DATA if row[0]!="SEC")
quote=sum(row[6] for row in DATA if row[0]!="SEC")
print("xlsx saved:",outx)
print(f"시장/원가추정합계={mkt:,} / 견적직접비합계={quote:,} / 예상직접이익={quote-mkt:,} ({round((quote-mkt)/mkt*100)}%)")

# HTML for PDF
rows_html=[]; no=0
for row in DATA:
    if row[0]=="SEC":
        rows_html.append(f'<tr class="sec"><td colspan="10">{row[1]}</td></tr>'); continue
    no+=1
    gubun,name,spec,model,mkt_u,conf,quote_u,note=row
    diff=quote_u-mkt_u
    marg="" if quote_u==0 else f"{round(diff/mkt_u*100)}%"
    dcls="neg" if diff<0 else ""
    rows_html.append(
        f'<tr><td class="c">{no}</td><td class="c">{gubun}</td><td class="nm">{name}</td>'
        f'<td class="sp">{spec}</td><td class="md">{model}</td>'
        f'<td class="r">{mkt_u:,}</td><td class="c">{conf}</td>'
        f'<td class="r">{quote_u:,}</td><td class="r {dcls}">{diff:,}</td><td class="r {dcls}">{marg}</td></tr>')
tot=f'<tr class="tot"><td colspan="5" class="tl">합계 — 시장/원가 추정 vs 견적 직접비</td><td class="r">{mkt:,}</td><td></td><td class="r">{quote:,}</td><td class="r">{quote-mkt:,}</td><td class="r">{round((quote-mkt)/mkt*100)}%</td></tr>'

html=f"""<!DOCTYPE html><html lang=ko><head><meta charset=UTF-8><style>
@page{{size:A4 landscape;margin:10mm 9mm;}}*{{margin:0;padding:0;box-sizing:border-box;}}
body{{font-family:"Malgun Gothic","맑은 고딕",sans-serif;color:#1f2937;font-size:8pt;-webkit-print-color-adjust:exact;print-color-adjust:exact;}}
:root{{--navy:#12315e;--red:#b91c1c;--sec:#dce6f5;--line:#c4cede;--gray:#6b7280;--yel:#fde68a;}}
.banner{{background:#fef2f2;border:1.5px solid var(--red);border-radius:6px;padding:7px 12px;margin-bottom:7px;}}
.banner b{{color:var(--red);font-size:12pt;}} .banner span{{color:#7f1d1d;font-size:8.3pt;}}
h1{{font-size:13pt;color:var(--navy);}}
.hd{{display:flex;justify-content:space-between;align-items:flex-end;border-bottom:2px solid var(--navy);padding-bottom:5px;margin-bottom:7px;}}
.logo{{font-size:16pt;font-weight:900;color:#1E40AF;}}
table{{width:100%;border-collapse:collapse;font-size:7.6pt;}}
th{{background:var(--navy);color:#fff;border:1px solid var(--navy);padding:4px 3px;text-align:center;}}
td{{border:1px solid var(--line);padding:3px 5px;vertical-align:middle;}}
td.c{{text-align:center;}}td.r{{text-align:right;}}td.nm{{font-weight:700;color:var(--navy);}}
td.sp{{font-size:7pt;}}td.md{{font-size:7.2pt;color:var(--navy);font-weight:600;}}
td.neg{{color:var(--red);font-weight:700;}}
tr.sec td{{background:var(--sec);font-weight:800;color:var(--navy);font-size:8pt;}}
tr.tot td{{background:var(--yel);font-weight:800;color:var(--navy);font-size:8.4pt;}}tr.tot td.tl{{text-align:right;}}
.note{{margin-top:8px;font-size:7.6pt;line-height:1.5;color:#374151;}}.note b{{color:var(--navy);}}.note .w{{color:var(--red);font-weight:700;}}
</style></head><body>
<div class=hd><div class=logo>UTTEC</div><h1>광안 보안등 조명제어 — 구매참고 BOM (실모델·시장가·마진)</h1><div style="font-size:8pt;color:var(--gray)">v1.0 · 2026-07-24</div></div>
<div class=banner><b>[내부 관리용 · 대외비]</b> <span>외부 제출 금지. 견적서(대외용) 원가·마진 검토 및 발주 참고용. 전용기기(MRBAS ELC)는 발주 전 ㈜엠알바스 실견적 필수.</span></div>
<table><thead><tr><th>No</th><th>구분</th><th>품목</th><th>도면 사양</th><th>권장 제조사·모델</th><th>시장단가<br>(추정)</th><th>신뢰</th><th>견적단가</th><th>차이</th><th>마진%</th></tr></thead>
<tbody>{chr(10).join(rows_html)}{tot}</tbody></table>
<div class=note>
<b>신뢰도</b>: ● 웹 실검색 확인 / ◐ 시장지식 추정 / ○ 전용기기(제조사 실견적 필수) / ▷ 용역(공임 원가추정) &nbsp;|&nbsp; <b>예상 직접이익</b> = 견적직접비 {quote:,} − 원가추정 {mkt:,} = <b>{quote-mkt:,}원 ({round((quote-mkt)/mkt*100)}%)</b>. 견적서엔 일반관리비 5%+이윤 10% 별도 가산.<br>
<span class=w>⚠ No.11 관제PC</span>: 브랜드 i5·32GB·1TB SSD 시세 > 견적단가 → 견적 상향 또는 사양 하향 검토. &nbsp; <span class=w>⚠ 전용기기</span>(SCU·0-10V 디밍모듈·릴레이 6eRM-w/16a·RPWR·HMI): 추정가 → 반드시 MRBAS 실견적 대체.<br>
<b>출처(2026-07-24 웹)</b>: mrbas.co.kr(ELC SYSTEM 카탈로그) · LS ELECTRIC/풍림·2sk·speedmall(차단기 ABS/EBS) · koreacable.net(F-CVVS) · HP/Dell 코리아(PC) · topbox·boxpara(함체).
</div></body></html>"""
outh="[내부용]삼환_광안보안등_조명제어_구매참고BOM_UTTEC_v1_20260724.html"
open(outh,"w",encoding="utf-8").write(html)
print("html saved:",outh)
